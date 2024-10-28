import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

class TreeNode:
    def __init__(self):
        self.children = {}
        self.records = []

class TitleTree:
    def __init__(self):
        self.root = TreeNode()

    def insert(self, key, record):
        """Insertar un registro en el árbol, agrupando por clave (ISRC + Identificadores)"""
        node = self.root
        key = str(key)
        for char in key:
            if char not in node.children:
                node.children[char] = TreeNode()
            node = node.children[char]
        node.records.append(record)

    def search(self, key):
        """Buscar todas las coincidencias de una clave en el árbol"""
        node = self.root
        key = str(key)
        for char in key:
            if char not in node.children:
                return None
            node = node.children[char]
        return node.records if node.records else None

def cargar_archivo_excel():
    archivos_excel = [f for f in os.listdir() if f.endswith('.xlsx')]
    
    print("Seleccione un archivo de la lista o escriba la ruta completa:")
    for i, archivo in enumerate(archivos_excel, 1):
        print(f"{i}. {archivo}")
    print("0. Escribir la ruta de otro archivo")

    opcion = input("Ingrese el número de su elección: ")
    try:
        opcion = int(opcion)
        if opcion == 0:
            archivo_seleccionado = input("Escriba la ruta completa del archivo: ")
        elif 1 <= opcion <= len(archivos_excel):
            archivo_seleccionado = archivos_excel[opcion - 1]
        else:
            print("Opción inválida.")
            return None
    except ValueError:
        print("Entrada inválida.")
        return None

    sheet_name = input("Ingrese el nombre de la hoja (deje en blanco para la primera hoja): ") or 0
    header_row = input("Ingrese la fila de encabezados (por defecto 1): ")
    header_row = int(header_row) - 1 if header_row.isdigit() else 0

    try:
        df = pd.read_excel(archivo_seleccionado, sheet_name=sheet_name, header=header_row)
        print(f"Archivo '{archivo_seleccionado}' cargado correctamente.")
        return df
    except Exception as e:
        print(f"Error al cargar el archivo '{archivo_seleccionado}': {e}")
        return None

def es_identificador_valido(identificador):
    return identificador and isinstance(identificador, str) and len(identificador) >= 6

def exportar_resultados(df_unificacion, df_agrupacion, df_no_agrupados):
    # Mover "Grupo Color" a la última posición en "Unificación_Obras"
    if "Grupo Color" in df_unificacion.columns:
        cols = [col for col in df_unificacion.columns if col != "Grupo Color"] + ["Grupo Color"]
        df_unificacion = df_unificacion[cols]

    with pd.ExcelWriter('100_Agrup_Unif.xlsx', engine='openpyxl', mode='w') as writer:
        df_unificacion.to_excel(writer, sheet_name='Unificación_Obras', index=False)
        df_agrupacion.to_excel(writer, sheet_name='Agrupación_Obras', index=False)
        df_no_agrupados.to_excel(writer, sheet_name='No Agrupados', index=False)

    wb = load_workbook('100_Agrup_Unif.xlsx')
    aplicar_colores(wb, 'Unificación_Obras', verde=True)
    aplicar_colores(wb, 'Agrupación_Obras')
    wb.save('100_Agrup_Unif.xlsx')
    print("Formato aplicado correctamente en las hojas.")

def aplicar_colores(wb, sheet_name, verde=False):
    ws = wb[sheet_name]
    col_color = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Grupo Color":
            col_color = col
            break

    if col_color is None:
        print(f"'Grupo Color' no encontrado en {sheet_name}.")
        return

    for i in range(2, ws.max_row + 1):
        color_hex = ws.cell(row=i, column=col_color).value
        if color_hex:
            color_fill = PatternFill(start_color="A8E6CF" if verde else color_hex, end_color="A8E6CF" if verde else color_hex, fill_type="solid")
            for j in range(1, ws.max_column + 1):
                ws.cell(row=i, column=j).fill = color_fill

def main():
    df_colaboraciones = cargar_archivo_excel()

    if df_colaboraciones is None:
        print("Error al cargar el archivo de Excel. Intente nuevamente.")
        return

    buscar_y_sumar_por_identificadores(df_colaboraciones)

def buscar_y_sumar_por_identificadores(df_colaboraciones):
    grupos_resultados = []
    obras_grupos = []
    obras_no_agrupadas = []

    invalid_values = ["N/A", "", "NO MLC", "NO", None]
    columnas_identificadores = [
        'MLC', 'ISWC', 'USA (BMI-ASCAP)', 'WORK ID', 'Harry Fox', 
        'MEXICO (SACM)', 'GUATEMALA (AEI)', 'COLOMBIA (SAYCO)', 'ACINPRO analogo', 
        'ACINPRO digital', 'ARGENTINA (SADAIC)', 'BRASIL', 'ESPAÑA SGAE'
    ]

    colores = ['A8E6CF', 'D6EDAF', 'B5D99C', 'CEE3A3', 'E2F0A8']
    color_index = 0

    tree = TitleTree()
    for _, row in df_colaboraciones.iterrows():
        # Crear una clave de agrupación basada en ISRC, ISWC y demás identificadores, ignorando lanzamiento y título
        key = tuple(row[col] for col in ['ISRC'] + columnas_identificadores if pd.notna(row[col]) and row[col] not in invalid_values)
        if key:
            tree.insert(key, row)

    for index, row in df_colaboraciones.iterrows():
        # Crear la misma clave para buscar coincidencias
        key = tuple(row[col] for col in ['ISRC'] + columnas_identificadores if pd.notna(row[col]) and row[col] not in invalid_values)

        if key:
            coincidencias = tree.search(key)
            if coincidencias:
                percentages_metadata = [match['%'] for match in coincidencias]
                total_suma = sum(percentages_metadata)

                nombres_completos_unificados = ', '.join(
                    (match['Autor'] + ' ' + match['Apellido']).strip() 
                    for match in coincidencias 
                    if pd.notna(match['Autor']) and pd.notna(match['Apellido'])
                )

                identificadores = {}
                for col in columnas_identificadores:
                    identificadores[col] = next(
                        (match[col] for match in coincidencias if pd.notna(match[col]) and match[col] not in invalid_values), '')

                grupo = {
                    'ISRC': row['ISRC'],
                    'Titulo': row.get('Titulo', ''),
                    'Autor': nombres_completos_unificados,
                    '%': total_suma,
                    'Grupo Color': colores[color_index % len(colores)]
                }
                grupo.update(identificadores)
                grupos_resultados.append(grupo)

                for obra in coincidencias:
                    obra['Grupo Color'] = colores[color_index % len(colores)]
                    obras_grupos.append(obra)

                color_index += 1
            else:
                row['Motivo'] = "Sin coincidencia por identificadores completos"
                obras_no_agrupadas.append(row)

    exportar_resultados(pd.DataFrame(grupos_resultados), pd.DataFrame(obras_grupos), pd.DataFrame(obras_no_agrupadas))


if __name__ == '__main__':
    titulo = "Agrupación y Unificación de obras"
    autor = "Alioth Musule A."

    color_rojo = "\033[31m"
    color_verde = "\033[32m"
    color_reset = "\033[0m"

    print(f"{color_rojo}╔═════════════════════════════════════════╗{color_reset}")
    print(f"{color_rojo}║{color_verde}  ██████████████████████████████████████ {color_reset}")
    print(f"{color_rojo}║{color_verde}  ██  {color_reset}{titulo} ")
    print(f"{color_rojo}║{color_verde}  ██████████████████████████████████████ {color_reset}")
    print(f"{color_rojo}╚═════════════════════════════════════════╝{color_reset}")
    print(f"                                Hecho por {autor}")
    
    main()  # Usa el archivo predeterminado
