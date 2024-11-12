import pandas as pd
import logging
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configuración del logging para un solo archivo de log
log_filename = "(ISWC)BusquedaAutor.log"
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(message)s', encoding='utf-8')

# Agregar un encabezado al inicio de cada ejecución
with open(log_filename, 'a', encoding='utf-8') as log_file:
    log_file.write(f"\n--- Ejecución iniciada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n")

# Cargar el archivo de datos unificados
archivo_origen = 'METADATA_PUBLISHING_UNIFICADO.xlsx'
archivo_destino = 'autor_busqueda.xlsx'

# Función para buscar y exportar registros por Autor y Apellido
def buscar_por_autor(df, autor, apellido):
    # Filtrar registros que contengan el nombre y apellido en las columnas Author y Last Name
    df_filtrado = df[(df['Author'].str.contains(autor, case=False, na=False)) & 
                     (df['Last Name'].str.contains(apellido, case=False, na=False))]
    
    # Nombre de la hoja será el nombre completo del autor
    nombre_hoja = f"{autor} {apellido}"
    
    # Si el nombre de hoja es demasiado largo (>31 caracteres), lo truncamos
    if len(nombre_hoja) > 31:
        nombre_hoja = nombre_hoja[:31]
    
    # Verificar si el archivo existe
    if os.path.exists(archivo_destino):
        # Cargar el workbook existente
        book = load_workbook(archivo_destino)
        writer = pd.ExcelWriter(archivo_destino, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        writer.book = book

        # Si la hoja existe, la eliminamos
        if nombre_hoja in writer.book.sheetnames:
            del writer.book[nombre_hoja]
        
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        df_filtrado.to_excel(writer, index=False, sheet_name=nombre_hoja)
        writer.save()
        writer.close()
    else:
        # Crear un nuevo archivo
        with pd.ExcelWriter(archivo_destino, engine='openpyxl') as writer:
            df_filtrado.to_excel(writer, index=False, sheet_name=nombre_hoja)
    
    # Aplicar colores a las filas en la hoja
    colores_visibles = [
        "FFEBCC", "FFF2CC", "E6FFCC", "CCFFE6", "CCE6FF", "E6CCFF", "FFD1DC", "FFCCCC",
        "CCFFEB", "CCE5FF", "E0FFE6", "FFFFCC", "FFCCE5", "FFEECC", "D6E6FF", "E6FFFA",
        "FFDAB9", "FFDFC4", "E6F7FF", "FFF9CC", "FFF1E0", "E8FFCC"
    ]
    
    # Abrir el workbook para aplicar los colores
    wb = load_workbook(archivo_destino)
    ws = wb[nombre_hoja]
    
    # Aplicar colores cíclicamente para diferenciar las filas
    for row_num in range(2, ws.max_row + 1):  # Empezamos desde la fila 2 para omitir el encabezado
        color = colores_visibles[(row_num - 2) % len(colores_visibles)]
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for col_num in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_num).fill = fill
    
    wb.save(archivo_destino)
    wb.close()
    
    # Log de la búsqueda realizada
    logging.info(f"Búsqueda completada para Autor: '{autor}' y Apellido: '{apellido}'.")
    logging.info(f"{len(df_filtrado)} registros encontrados y exportados a '{archivo_destino}' en la hoja '{nombre_hoja}'.")
    
    print(f"Se encontraron y exportaron {len(df_filtrado)} registros con Autor '{autor}' y Apellido '{apellido}'.")

# Cargar los datos desde la hoja 'Unificados'
df = pd.read_excel(archivo_origen, sheet_name='Unificados')

# Ingresar el nombre del Autor y Apellido para realizar la búsqueda
autor = input("Ingresa el nombre del Autor: ")
apellido = input("Ingresa el Apellido: ")

# Realizar la búsqueda y exportar los registros
buscar_por_autor(df, autor, apellido)
