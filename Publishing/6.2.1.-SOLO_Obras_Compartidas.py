import pandas as pd
import logging
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configuración del logging para un solo archivo de log
log_filename = "(ISWC)BusquedaColaboraciones.log"
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(message)s', encoding='utf-8')

# Agregar un encabezado al inicio de cada ejecución
with open(log_filename, 'a', encoding='utf-8') as log_file:
    log_file.write(f"\n--- Ejecución iniciada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n")

# Cargar el archivo de datos unificados
archivo_origen = 'METADATA_PUBLISHING_UNIFICADO.xlsx'
archivo_colaboraciones = 'METADATA_U_RECOPILACION_Colaboraciones.xlsx'

# Función para buscar y exportar registros de colaboraciones
def buscar_colaboraciones(df):
    # Filtrar registros que tengan colaboraciones (verificar comas o puntos en 'Author' o 'Last Name')
    colaboraciones = df[df['Author'].str.contains(r'[•,]', na=False) | df['Last Name'].str.contains(r'[•,]', na=False)]
    
    if colaboraciones.empty:
        print("No se encontraron registros de colaboraciones.")
        return

    # Verificar si el archivo de colaboraciones ya existe y cargar datos existentes
    if os.path.exists(archivo_colaboraciones):
        # Cargar el archivo y la hoja existente para agregar nuevas colaboraciones
        with pd.ExcelWriter(archivo_colaboraciones, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Leer el archivo existente y combinarlo con las nuevas colaboraciones
            print("leyendo archivo existente")
            wb = load_workbook(archivo_colaboraciones)
            ws = wb["Colaboraciones"]
            datos_existentes = pd.read_excel(archivo_colaboraciones, sheet_name="Colaboraciones")
            datos_completos = pd.concat([datos_existentes, colaboraciones], ignore_index=True)
            # Sobreescribir todos los datos combinados en el archivo
            datos_completos.to_excel(writer, index=False, sheet_name="Colaboraciones")
            accion = "actualizado"
    else:
        # Si el archivo no existe, crear uno nuevo y guardar las colaboraciones
        with pd.ExcelWriter(archivo_colaboraciones, engine='openpyxl') as writer:
            print("creando archivo")
            colaboraciones.to_excel(writer, index=False, sheet_name="Colaboraciones")
            accion = "creado"

    # Aplicar colores alternos a todas las filas
    wb = load_workbook(archivo_colaboraciones)
    ws = wb["Colaboraciones"]
    colores_visibles = [
        "FFEBCC", "FFF2CC", "E6FFCC", "CCFFE6", "CCE6FF", "E6CCFF", "FFD1DC", "FFCCCC",
        "CCFFEB", "CCE5FF", "E0FFE6", "FFFFCC", "FFCCE5", "FFEECC", "D6E6FF", "E6FFFA",
        "FFDAB9", "FFDFC4", "E6F7FF", "FFF9CC", "FFF1E0", "E8FFCC"
    ]
    
    # Aplicar colores a todas las filas, incluyendo las nuevas
    for row_num in range(2, ws.max_row + 1):  # Empezamos desde la fila 2 para omitir el encabezado
        color = colores_visibles[(row_num - 2) % len(colores_visibles)]
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for col_num in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_num).fill = fill
    
    wb.save(archivo_colaboraciones)
    wb.close()
    
    # Log de la búsqueda de colaboraciones
    logging.info(f"Archivo de colaboraciones '{archivo_colaboraciones}' {accion} con {len(colaboraciones)} registros nuevos.")
    print(f"Archivo de colaboraciones '{archivo_colaboraciones}' {accion} exitosamente con {len(colaboraciones)} registros nuevos.")

# Cargar los datos desde la hoja 'Unificados'
df = pd.read_excel(archivo_origen, sheet_name='Unificados')

# Realizar la búsqueda y exportar las colaboraciones
buscar_colaboraciones(df)
