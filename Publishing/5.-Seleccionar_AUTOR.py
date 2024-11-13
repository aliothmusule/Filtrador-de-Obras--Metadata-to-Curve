import pandas as pd
import logging
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configuración del logging para un solo archivo de log
log_filename = "(ISWC)BusquedaAutor.log"
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(message)s', encoding='utf-8')

# Agregar un encabezado al inicio de cada ejecución
with open(log_filename, 'a', encoding='utf-8') as log_file:
    log_file.write(f"\n--- Ejecución iniciada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---\n")

# Cargar el archivo de datos unificados
archivo_origen = 'METADATA_PUBLISHING_UNIFICADO.xlsx'
directorio_autores = 'Autores_INDCOLB'

# Asegurarse de que la carpeta 'Autores' exista
if not os.path.exists(directorio_autores):
    os.makedirs(directorio_autores)

# Función para buscar y exportar registros por Autor y Apellido
def buscar_por_autor(df, autor, apellido):
    # Filtrar registros que contengan el nombre y apellido en las columnas Author y Last Name
    posibles_autores = df[(df['Author'].str.contains(autor, case=False, na=False)) & 
                          (df['Last Name'].str.contains(apellido, case=False, na=False))]
    
    # Filtrar registros que tengan colaboraciones y contengan específicamente el autor y apellido seleccionados
    colaboraciones = posibles_autores[
        (posibles_autores['Author'].str.contains(r'[•,]', na=False) | posibles_autores['Last Name'].str.contains(r'[•,]', na=False)) &
        (posibles_autores['Author'].str.contains(autor, case=False, na=False)) &
        (posibles_autores['Last Name'].str.contains(apellido, case=False, na=False))
    ]
    
    # Si no se encontraron coincidencias
    if posibles_autores.empty:
        print(f"No se encontraron coincidencias para Autor '{autor}' y Apellido '{apellido}'.")
        return

    # Si hay varios autores coincidentes, pedir al usuario que seleccione uno
    if len(posibles_autores['Author'].unique()) > 1:
        print("Se encontraron varios autores coincidentes:")
        for idx, (nombre, apellido) in enumerate(posibles_autores[['Author', 'Last Name']].drop_duplicates().itertuples(index=False), start=1):
            print(f"{idx}. {nombre} {apellido}")
        seleccion = int(input("Seleccione el número del autor que desea exportar: ")) - 1
        autor_seleccionado, apellido_seleccionado = posibles_autores[['Author', 'Last Name']].drop_duplicates().iloc[seleccion]
        df_filtrado = df[(df['Author'] == autor_seleccionado) & (df['Last Name'] == apellido_seleccionado)]
    else:
        autor_seleccionado, apellido_seleccionado = autor, apellido
        df_filtrado = posibles_autores

    # Nombre del archivo será el nombre completo del autor seleccionado
    nombre_archivo = os.path.join(directorio_autores, f"{autor_seleccionado}_{apellido_seleccionado}.xlsx")
    archivo_existente = os.path.exists(nombre_archivo)
    
    # Guardar o actualizar el archivo Excel individual del autor
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl', mode='a' if archivo_existente else 'w') as writer:
        if archivo_existente:
            book = writer.book
            if "Registros" in book.sheetnames:
                del book["Registros"]
            if "Colaboraciones" in book.sheetnames:
                del book["Colaboraciones"]
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            df_filtrado.to_excel(writer, index=False, sheet_name="Registros")
            colaboraciones.to_excel(writer, index=False, sheet_name="Colaboraciones")
            accion = "actualizado"
        else:
            df_filtrado.to_excel(writer, index=False, sheet_name="Registros")
            colaboraciones.to_excel(writer, index=False, sheet_name="Colaboraciones")
            accion = "creado"

    # Aplicar colores a las filas en la hoja
    colores_visibles = [
        "FFEBCC", "FFF2CC", "E6FFCC", "CCFFE6", "CCE6FF", "E6CCFF", "FFD1DC", "FFCCCC",
        "CCFFEB", "CCE5FF", "E0FFE6", "FFFFCC", "FFCCE5", "FFEECC", "D6E6FF", "E6FFFA",
        "FFDAB9", "FFDFC4", "E6F7FF", "FFF9CC", "FFF1E0", "E8FFCC"
    ]
    
    # Aplicar colores cíclicamente en "Registros"
    wb = load_workbook(nombre_archivo)
    ws = wb["Registros"]
    for row_num in range(2, ws.max_row + 1):
        color = colores_visibles[(row_num - 2) % len(colores_visibles)]
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for col_num in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_num).fill = fill
    
    # Guardar y cerrar
    wb.save(nombre_archivo)
    wb.close()
    
    # Log de la búsqueda realizada
    logging.info(f"Búsqueda completada para Autor: '{autor_seleccionado}' y Apellido: '{apellido_seleccionado}'.")
    logging.info(f"{len(df_filtrado)} registros encontrados y {accion} en '{nombre_archivo}'.")
    logging.info(f"{len(colaboraciones)} registros de colaboraciones exportados en 'Colaboraciones'.")
    
    print(f"Se encontraron y exportaron {len(df_filtrado)} registros para '{autor_seleccionado} {apellido_seleccionado}'.")
    print(f"Archivo '{nombre_archivo}' {accion} exitosamente.")
    print(f"{len(colaboraciones)} registros de colaboraciones agregados a la hoja 'Colaboraciones'.")

# Cargar los datos desde la hoja 'Unificados'
df = pd.read_excel(archivo_origen, sheet_name='Unificados')

# Ingresar el nombre del Autor y Apellido para realizar la búsqueda
autor = input("Ingresa el nombre del Autor: ")
apellido = input("Ingresa el Apellido: ")

# Realizar la búsqueda y exportar los registros
buscar_por_autor(df, autor, apellido)
