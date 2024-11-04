import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter

# Definir las editoras a filtrar
editoras = [
    "BACKSTAGE EDITORA TX", "MUSIC BY BACKSTAGE PUBLISHING", "MUSIC BY CANZION LEGACY",
    "CANZION LEGACY", "CANZION PUBLISHING TX", "GRUPO CANZION EDITORA",
    "MUSIC BY HEAVEN LEGAZY", "HEAVEN LEGACY", "HEAVEN NETWORKS", "MUSIC BY HEAVEN PUBLISHING",
    "PUBLISHING BY COMPOSITORES", "COMPOSITORES PUBLISHING", "ALIENTO PUBLISHING",
    "CEV ADMINISTRATION", "CEV PUBLISHER LLC", "LK COLLECTIVE", "FRILOP MUSIC",
    "JUAN SALINAS MUSIC PUBLISHING", "LOS VENCEDORES PUBLISHING", "MUSIC BY GREEN MUSIC PUBLISHING",
    "MUSIC BY MAJESTIC RECORDS PUBLISHING", "MUSIC FROM AZ MUSIC", "REDIMI2 MUSIC PUBLISHING",
    "REYVOL MUSIK LLC", "COALO ZAMORANO PUBLISHING", "MIKE RODRIGUEZ MUSIC",
    "ADORA MUSIC INTERNATIONAL", "MAR DE CRISTAL PUBLISHING", "MONTESANTO PUBLISHING",
    "GRACE ESPANOL MUSICA"
]

# Colores más visibles para el fondo de cada grupo
colores_visibles = [
    "#FFDDC1", "#C1FFC1", "#C1D4FF", "#FFC1E3", "#D4C1FF", "#C1FFFF", "#FFE7C1",
    "#FFD1A1", "#DAFFC1", "#C1EAF0", "#E0C1FF", "#C1F0E8", "#FFD5D1", "#C1E0DA",
    "#F0EAC1", "#D1FFC1", "#C1D1FF", "#F2C1C1", "#DAD1A1", "#EAD1C1"
]


def asignar_color_grupo(grupo_id):
    return colores_visibles[(grupo_id - 1) % len(colores_visibles)]

class TreeNode:
    def __init__(self):
        self.children = {}
        self.records = []

class TitleTree:
    def __init__(self):
        self.root = TreeNode()

    def insert(self, key, record):
        node = self.root
        for part in key:
            if part not in node.children:
                node.children[part] = TreeNode()
            node = node.children[part]
        node.records.append(record)

    def get_groups(self):
        groups = []
        def traverse(node):
            if node.records:
                groups.append(node.records)
            for child in node.children.values():
                traverse(child)
        traverse(self.root)
        return groups

print("Cargando el archivo de metadatos...")
archivo_metadata = 'METADATA_PUBLISHING_COLOR.xlsx'
metadata_df = pd.read_excel(archivo_metadata, header=0)
print("Archivo cargado exitosamente.")
metadata_df['Publisher'] = metadata_df['Publisher'].astype(str).str.strip()

# Filtrar por editoras
metadata_publishing_df = metadata_df[metadata_df['Publisher'].isin(editoras)].copy()

columnas_agrupacion = ['Artista', 'Titulo', 'Album', 'ISRC', 'Lanzamiento']
metadata_publishing_df[columnas_agrupacion] = metadata_publishing_df[columnas_agrupacion].astype(str)
metadata_publishing_df['%'] = pd.to_numeric(metadata_publishing_df['%'], errors='coerce')

print("Insertando registros en el árbol...")
tree = TitleTree()
for _, row in metadata_publishing_df.iterrows():
    key = tuple(row[col] for col in columnas_agrupacion)
    tree.insert(key, row)

import math  # Importar la biblioteca math para verificar NaN

import math  # Importar la biblioteca math para verificar NaN

# Calcular el porcentaje de contrato, redondeado a 2 decimales, y asignar Grupo Contador a cada grupo
grupo_id = 1
for group in tree.get_groups():
    total_porcentaje = sum(float(record['%']) for record in group if pd.notna(record['%']))
    for index, record in enumerate(group):
        # Calcular el valor redondeado a 2 decimales, si el total es mayor a 0
        contrato_redondeado = round((float(record['%']) / total_porcentaje) * 100, 2) if total_porcentaje else 0
        
        # Verificar si contrato_redondeado es NaN antes de continuar
        if math.isnan(contrato_redondeado):
            print(f"Valor NaN encontrado en la fila {index + 1} del grupo {grupo_id}: {record}")
            record['Contrato'] = 0  # O el valor que desees asignar en caso de NaN
        else:
            # Verificar si el valor redondeado tiene un decimal exacto ".0"
            if contrato_redondeado == int(contrato_redondeado):
                record['Contrato'] = int(contrato_redondeado)  # Mostrar como entero si es, por ejemplo, 5.0
            else:
                record['Contrato'] = contrato_redondeado  # Mostrar tal cual si es, por ejemplo, 5.01, 5.10, 5.20
        
        record['Grupo Contador'] = grupo_id
    grupo_id += 1

# Crear DataFrame y clasificar por porcentaje total de contrato
metadata_publishing_df = pd.DataFrame([record for group in tree.get_groups() for record in group])
grupos_100 = metadata_publishing_df[metadata_publishing_df['Contrato'] == 100].copy()
grupos_menor_100 = metadata_publishing_df[metadata_publishing_df['Contrato'] < 100].copy()

archivo_exportado = 'METADATA_PUBLISHING_SEPARADO.xlsx'

if not os.path.exists(archivo_exportado):
    with pd.ExcelWriter(archivo_exportado, engine='xlsxwriter') as writer:
        for hoja, data in zip(['Grupos 100%', 'Grupos < 100%'], [grupos_100, grupos_menor_100]):
            data.to_excel(writer, index=False, sheet_name=hoja)
            workbook = writer.book
            worksheet = writer.sheets[hoja]
            for row_num, grupo in enumerate(data['Grupo Contador'], start=1):
                color = asignar_color_grupo(grupo)
                cell_format = workbook.add_format({'bg_color': color})
                worksheet.set_row(row_num, None, cell_format)
    print(f"Archivo exportado y coloreado exitosamente como {archivo_exportado}")

else:
    print("El archivo ya existe. Aplicando colores faltantes con openpyxl...")
    wb = load_workbook(archivo_exportado)
    for hoja, data in zip(['Grupos 100%', 'Grupos < 100%'], [grupos_100, grupos_menor_100]):
        ws = wb[hoja]
        for index, row in data.iterrows():
            row_num = index + 2
            color_hex = asignar_color_grupo(row['Grupo Contador'])
            fill = PatternFill(start_color=color_hex[1:], end_color=color_hex[1:], fill_type="solid")
            for cell in ws[row_num]:
                cell.fill = fill
    wb.save(archivo_exportado)
    print(f"Archivo existente actualizado exitosamente como {archivo_exportado}")
