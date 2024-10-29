import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter



# .---------------- CODIGO ACTUAL PARA SACAR LOS GRUPOS DE MISMAS OBRAS, COLOREAR LOS GRUPOS Y SACAR
# ----------------- LOS PORCENTAJES DE CONTRATO.

#---------------------------FALTARÍA : PODER JUNTAR LOS GRUPOS EN UNO SOLO PARA TENER UNIDOS TODOS LOS IDENTIFICADORES
#---------------------------EXISTENTES ENTRE ESOS GRUPOS Y ASÍ PODER EMPEZAR CON LA LOGICA DE AGREGAR O ACTUALIZAR
#.------------------------.-LOS DATOS DENTRO DE LA PLANTILLA DE CURVE.

# Definir las editoras a filtrar
editoras = [
    "BACKSTAGE EDITORA TX", "MUSIC BY BACKSTAGE PUBLISHING", "MUSIC BY CANZION LEGACY",
    "CANZION LEGACY", "CANZION PUBLISHING TX", "GRUPO CANZION EDITORA",
    "MUSIC BY HEAVEN LEGAZY", "HEAVEN LEGACY", "HEAVEN NETWORKS", "MUSIC BY HEAVEN PUBLISHING",
    "PUBLISHING BY COMPOSITORES", "COMPOSITORES PUBLISHING", "ALIENTO PUBLISHING",
    "CEV ADMINISTRATION", "CEV PUBLISHER LLC", "LK COLLECTIVE", "FRILOP MUSIC",
    "JUAN SALINAS MUSIC PUBLISHING", "LOS VENCEDORES PUBLISHING", "MUSIC BY GREEN MUSIC PUBLISHING",
    "MUSIC BY MAJESTIC RECORDS PUBLISHING", "MUSIC FROM AZ MUSIC", "REDIMI2 MUSIC PUBLISHING",
    "REYVOL MUSIK LLC", "COALO ZAMORANO PUBLISHING", "MIKE RODRIGUEZ MUSIC",
    "ADORA MUSIC INTERNATIONAL", "MAR DE CRISTAL PUBLISHING", "MONTESANTO PUBLISHING",
    "GRACE ESPANOL MUSICA"
]

# Colores más visibles para el fondo de cada grupo
colores_visibles = [
    "#FFD966",  # Amarillo claro
    "#93C47D",  # Verde pastel
    "#F6B26B",  # Naranja suave
    "#9FC5E8",  # Azul claro
    "#FFE599",  # Amarillo suave
    "#A2C4C9",  # Aqua pastel
    "#D5A6BD",  # Rosa suave
    "#D9EAD3",  # Verde muy claro
    "#C9DAF8",  # Azul celeste
    "#EAD1DC",  # Rosa pálido
    "#FFCCFF",  # Rosa lavanda
    "#CCCCFF",  # Azul lavanda
    "#FFCCCC",  # Rosa salmón claro
    "#D9C2E9",  # Lila claro
    "#CFE2F3",  # Celeste muy claro
    "#E2F0D9",  # Verde menta suave
    "#FFF2CC",  # Amarillo mantequilla
    "#F4CCCC",  # Rosa suave pastel
    "#E6B8AF",  # Rosa coral claro
    "#F9CB9C"   # Naranja melocotón claro
]


# Asignar colores alternos a cada grupo
def asignar_color_grupo(grupo_id):
    return colores_visibles[(grupo_id - 1) % len(colores_visibles)]

# Estructura del árbol para organizar los registros
class TreeNode:
    def __init__(self):
        self.children = {}
        self.records = []

class TitleTree:
    def __init__(self):
        self.root = TreeNode()

    def insert(self, key, record):
        node = self.root
        for part in key:
            if part not in node.children:
                node.children[part] = TreeNode()
            node = node.children[part]
        node.records.append(record)

    def get_groups(self):
        groups = []
        def traverse(node):
            if node.records:
                groups.append(node.records)
            for child in node.children.values():
                traverse(child)
        traverse(self.root)
        return groups

# Cargar el archivo METADATA CENTRAL.xlsx
print("Cargando el archivo de metadatos...")
archivo_metadata = 'METADATA CENTRAL.xlsx'
metadata_df = pd.read_excel(archivo_metadata, header=1)  # El encabezado está en la fila 2
print("Archivo cargado exitosamente.")
metadata_df['Publisher'] = metadata_df['Publisher'].astype(str).str.strip()

# Filtrar las filas donde el "Publisher" esté en la lista de editoras
metadata_publishing_df = metadata_df[metadata_df['Publisher'].isin(editoras)].copy()

columnas_agrupacion = ['Artista', 'Titulo', 'Album', 'ISRC', 'Lanzamiento']
metadata_publishing_df[columnas_agrupacion] = metadata_publishing_df[columnas_agrupacion].astype(str)

# Asegurarse de que la columna "%" esté en formato numérico
metadata_publishing_df['%'] = pd.to_numeric(metadata_publishing_df['%'], errors='coerce')

# Crear el árbol e insertar registros basados en la clave de agrupación
print("Insertando registros en el árbol...")
tree = TitleTree()
for _, row in metadata_publishing_df.iterrows():
    key = tuple(row[col] for col in columnas_agrupacion)  # Clave de agrupación
    tree.insert(key, row)

# Calcular el porcentaje de contrato y asignar Grupo Contador a cada grupo
grupo_id = 1
for group in tree.get_groups():
    total_porcentaje = sum(float(record['%']) for record in group if pd.notna(record['%']))
    for record in group:
        record['Contrato'] = (float(record['%']) / total_porcentaje) * 100 if total_porcentaje else 0
        record['Grupo Contador'] = grupo_id
    grupo_id += 1

# Convertir los datos de nuevo a DataFrame con el contrato y contador de grupo asignado
metadata_publishing_df = pd.DataFrame([record for group in tree.get_groups() for record in group])

# Reordenar las columnas para que "Contrato" esté inmediatamente después de "%"
cols = metadata_publishing_df.columns.tolist()
indice_porcentaje = cols.index('%')
cols.insert(indice_porcentaje + 1, cols.pop(cols.index('Contrato')))
metadata_publishing_df = metadata_publishing_df[cols]

# Nombre del archivo exportado
archivo_exportado = 'METADATA_PUBLISHING_COLOR.xlsx'

# Verificar si el archivo ya existe
if not os.path.exists(archivo_exportado):
    # Si el archivo no existe, crearlo con xlsxwriter y aplicar los colores
    with pd.ExcelWriter(archivo_exportado, engine='xlsxwriter') as writer:
        metadata_publishing_df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Aplicar colores por grupo de manera rápida
        for row_num, grupo in enumerate(metadata_publishing_df['Grupo Contador'], start=1):  # Empieza desde la fila 1
            color = asignar_color_grupo(grupo)
            cell_format = workbook.add_format({'bg_color': color})
            worksheet.set_row(row_num, None, cell_format)  # Aplica el color a toda la fila

    print(f"Archivo exportado y coloreado exitosamente como {archivo_exportado}")

else:
    # Si el archivo ya existe, abrirlo con openpyxl y aplicar los colores faltantes
    print("El archivo ya existe. Aplicando colores faltantes con openpyxl...")
    wb = load_workbook(archivo_exportado)
    ws = wb.active

    # Aplicar colores usando openpyxl en base a "Grupo Contador"
    for index, row in metadata_publishing_df.iterrows():
        row_num = index + 2  # Ajustar el índice para la fila en Excel
        color_hex = asignar_color_grupo(row['Grupo Contador'])
        fill = PatternFill(start_color=color_hex[1:], end_color=color_hex[1:], fill_type="solid")

        for cell in ws[row_num]:  # Aplicar el color a toda la fila
            cell.fill = fill

    wb.save(archivo_exportado)
    print(f"Archivo existente actualizado exitosamente como {archivo_exportado}")
